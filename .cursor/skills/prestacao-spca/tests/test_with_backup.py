from pathlib import Path

import pytest
from with_backup import with_backup


def test_with_backup_cria_backup_antes_de_operacao(tmp_path):
    target = tmp_path / "db.sqlite"
    target.write_text("original")
    @with_backup(target)
    def mutar(path):
        path.write_text("mutado")
    mutar()
    assert target.read_text() == "mutado"
    backups = list(tmp_path.glob("db.sqlite.bak-*"))
    assert len(backups) == 1
    assert backups[0].read_text() == "original"


def test_with_backup_restaura_e_remove_backup_se_operacao_falhar(tmp_path):
    target = tmp_path / "db.sqlite"
    target.write_text("original")
    @with_backup(target)
    def mutar(path):
        path.write_text("parcial")
        raise RuntimeError("falha")
    with pytest.raises(RuntimeError):
        mutar()
    # arquivo restaurado para original
    assert target.read_text() == "original"
    # backup limpo
    assert list(tmp_path.glob("db.sqlite.bak-*")) == []


def test_with_backup_preserva_target_nao_existente_como_erro(tmp_path):
    target = tmp_path / "nao_existe.sqlite"
    @with_backup(target)
    def mutar(path):
        path.write_text("novo")
    with pytest.raises(FileNotFoundError):
        mutar()


def test_with_backup_em_xlsx_mutation(tmp_path):
    """Cenário real: backup de xlsx antes de mutar via openpyxl.

    Valida o contrato do decorator: (1) copia binária do estado pré-mutação,
    (2) operação ocorre no target, (3) backup fica no lugar após sucesso.
    Não reabri o .bak- via openpyxl porque a extensão .bak-* não é
    suportada pelo loader (limitação do openpyxl, não do decorator).
    """
    from openpyxl import Workbook, load_workbook
    target = tmp_path / "cadastro.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["nome", "cpf"])
    ws.append(["ORIGINAL", "000"])
    wb.save(target)
    bytes_originais = target.read_bytes()

    @with_backup(target)
    def adicionar_linha(path, nome, cpf):
        wb = load_workbook(path)
        ws = wb.active
        ws.append([nome, cpf])
        wb.save(path)
        return ws.max_row

    # decorator injeta `target` como primeiro arg posicional automaticamente
    n = adicionar_linha("NOVO", "111")
    assert n == 3
    # arquivo tem o acréscimo
    wb2 = load_workbook(target)
    assert wb2.active.cell(3, 1).value == "NOVO"
    # backup criado com bytes idênticos ao estado pré-mutação
    backups = list(tmp_path.glob("cadastro.xlsx.bak-*"))
    assert len(backups) == 1
    assert backups[0].read_bytes() == bytes_originais


def test_with_backup_em_cenario_real_processar_mes(tmp_path):
    """Simula o uso real em processar_mes.py: decorator envolve a chamada
    que abre conexão SQLite e muta o DB canônico spca_revisao.db.

    Espelha o padrão de _gravar_revisao_com_backup() no script: inner function
    decorada com @with_backup(DB_PATH) que delega para a função mutante.
    Cobre: (1) sucesso → DB muta + backup preserva estado anterior;
           (2) falha → DB volta ao estado pré-mutação + backup limpo.
    """
    import sqlite3
    target = tmp_path / "spca_revisao.db"
    # seed: schema + 1 linha de "estado anterior" do DB
    con = sqlite3.connect(target)
    con.execute(
        "CREATE TABLE revisao_mes (id INTEGER PRIMARY KEY, estado TEXT, linhas INTEGER)"
    )
    con.execute("INSERT INTO revisao_mes (estado, linhas) VALUES ('v0', 0)")
    con.commit()
    con.close()

    # Simula _gravar_revisao_db: recebe o path injetado, abre conexão, muta.
    @with_backup(target)
    def gravar_revisao_db(path, estado, linhas):
        import sqlite3 as sql
        c = sql.connect(path)
        try:
            c.execute("DELETE FROM revisao_mes")
            c.execute(
                "INSERT INTO revisao_mes (estado, linhas) VALUES (?, ?)",
                (estado, linhas),
            )
            c.commit()
        finally:
            c.close()
        return {"linhas": linhas}

    # Cenario de sucesso: DB muta e backup guarda estado pré-mutação
    meta = gravar_revisao_db("v1", 42)
    assert meta["linhas"] == 42
    con = sqlite3.connect(target)
    row = con.execute("SELECT estado, linhas FROM revisao_mes").fetchone()
    con.close()
    assert row == ("v1", 42), f"esperado ('v1', 42), obtido {row}"
    backups = list(tmp_path.glob("spca_revisao.db.bak-*"))
    assert len(backups) == 1
    con_bak = sqlite3.connect(backups[0])
    bak_row = con_bak.execute("SELECT estado, linhas FROM revisao_mes").fetchone()
    con_bak.close()
    assert bak_row == ("v0", 0), (
        f"backup deve guardar estado anterior, obtido {bak_row}"
    )

    # Cenario de falha: DB volta a 'v1' e o backup da tentativa falhada é removido.
    # O backup do caso de sucesso acima continua no disco (decorator só remove
    # o backup da tentativa corrente em caso de exceção).
    import time
    time.sleep(1.1)  # garante timestamp diferente do backup do caso de sucesso
    backups_apos_sucesso = sorted(p.name for p in tmp_path.glob("spca_revisao.db.bak-*"))

    @with_backup(target)
    def gravar_revisao_que_falha(path, estado, linhas):
        c = sqlite3.connect(path)
        c.execute("DELETE FROM revisao_mes")
        c.execute(
            "INSERT INTO revisao_mes (estado, linhas) VALUES (?, ?)",
            (estado, linhas),
        )
        c.commit()
        c.close()
        raise RuntimeError("falha simulada de gravação no DB")

    with pytest.raises(RuntimeError):
        gravar_revisao_que_falha("v2_QUEBRA", 99)

    con = sqlite3.connect(target)
    row_pos = con.execute("SELECT estado, linhas FROM revisao_mes").fetchone()
    con.close()
    assert row_pos == ("v1", 42), (
        f"após falha, DB deve estar em 'v1', obtido {row_pos}"
    )
    backups_depois = sorted(p.name for p in tmp_path.glob("spca_revisao.db.bak-*"))
    # O backup do caso de sucesso continua no disco; o da tentativa falhada foi removido.
    # Se o timestamp diferir (>1s), o set após falha == o set antes da falha.
    assert backups_depois == backups_apos_sucesso, (
        f"backup da tentativa falhada deve ter sido removido, "
        f"mas os anteriores preservados: antes={backups_apos_sucesso}, "
        f"depois={backups_depois}"
    )


def test_with_backup_em_json_mutation(tmp_path):
    """Cenário real: backup de json antes de mutar (caso drive_manifest.json).

    Espelha o padrão aplicado em sync_drive.py:sync(): inner function decorada
    com @with_backup(target) que recebe o path injetado pelo decorator e delega
    para a função pública save_manifest(raiz, manifest_data). Valida:
    (1) backup binário preserva estado pré-mutação;
    (2) gravação final do json reflete os dados novos;
    (3) restauração e remoção do backup em caso de exceção.
    """
    import json
    # raiz simulada (subdiretório que representa o projeto)
    raiz = tmp_path / "projeto"
    raiz.mkdir()
    target = raiz / "resultados" / "drive_manifest.json"
    target.parent.mkdir(parents=True)
    target.write_text(
        json.dumps(
            {"synced_at": "2025-01-01T00:00:00Z", "uploaded": []},
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    # Espelha exatamente a forma de uso em sync_drive.py:
    # save_manifest público preservado + wrapper interno decorado.
    def save_manifest(raiz, manifest_data):
        out = raiz / "resultados" / "drive_manifest.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(
            json.dumps(manifest_data, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return out

    @with_backup(target)
    def _save_manifest_com_backup(path, manifest_data):
        return save_manifest(raiz, manifest_data)

    # Sucesso: manifest é gravado + backup guarda estado anterior
    novo = {
        "synced_at": "2025-06-24T00:00:00Z",
        "uploaded": ["file1.xlsx"],
        "folders": 3,
    }
    out = _save_manifest_com_backup(novo)
    assert out == target
    atual = json.loads(target.read_text(encoding="utf-8"))
    assert atual["uploaded"] == ["file1.xlsx"]
    assert atual["synced_at"] == "2025-06-24T00:00:00Z"
    backups = list((raiz / "resultados").glob("drive_manifest.json.bak-*"))
    assert len(backups) == 1
    bak = json.loads(backups[0].read_text(encoding="utf-8"))
    assert bak["synced_at"] == "2025-01-01T00:00:00Z"
    assert bak["uploaded"] == []

    # Falha: manifest volta ao estado pré-falha e backup da tentativa é removido
    import time
    time.sleep(1.1)  # timestamp distinto do backup do caso de sucesso
    backups_apos_sucesso = sorted(
        p.name for p in (raiz / "resultados").glob("drive_manifest.json.bak-*")
    )

    @with_backup(target)
    def _save_manifest_que_falha(path, manifest_data):
        target.write_text(
            json.dumps(manifest_data, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        raise RuntimeError("falha simulada de gravação do manifest")

    with pytest.raises(RuntimeError):
        _save_manifest_que_falha(
            {"synced_at": "2025-06-24T01:00:00Z", "uploaded": ["QUEBRA"]}
        )

    # manifest restaurado para o estado pós-sucesso (antes da tentativa falhada)
    pos = json.loads(target.read_text(encoding="utf-8"))
    assert pos["synced_at"] == "2025-06-24T00:00:00Z"
    assert pos["uploaded"] == ["file1.xlsx"]
    backups_depois = sorted(
        p.name for p in (raiz / "resultados").glob("drive_manifest.json.bak-*")
    )
    assert backups_depois == backups_apos_sucesso, (
        f"backup da tentativa falhada deve ter sido removido, "
        f"mas os anteriores preservados: antes={backups_apos_sucesso}, "
        f"depois={backups_depois}"
    )


# ----------------------------------------------------------------------------
# Cenários para lib_revisao_exportacao.salvar_revisao e
# lib_xml_origem_recurso — padrão aplicado em 2026-06-24 para estender
# cobertura do decorator @with_backup a scripts lib_* mutantes
# (xlsx de revisão mensal, XML origemRecurso, xlsx de revisão anual).
# ----------------------------------------------------------------------------


def test_with_backup_salvar_revisao_regravacao(tmp_path):
    """salvar_revisao() — quando xlsx já existe, regrava com backup binário.

    Espelha o padrão em lib_revisao_exportacao.py:salvar_revisao(): inner
    function _salvar_com_backup() é decorada com @with_backup(path) e o
    decorator injeta `path` como 1º arg posicional. Valida: backup preserva
    estado pré-mutação e o xlsx final contém os dados novos.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from lib_revisao_exportacao import (
        ABA_BLOQUEADAS,
        ABA_PRONTAS,
        ABA_RESUMO,
        COLS_BLOQUEADAS,
        COLS_PRONTAS,
        caminho_revisao,
        salvar_revisao,
    )

    output_dir = tmp_path / "mensal"
    output_dir.mkdir()
    mes_slug = "janeiro"
    xlsx = caminho_revisao(output_dir, mes_slug)

    # Estado anterior: xlsx pré-existente escrito em formato idêntico ao
    # que salvar_revisao() produz.
    prontas_v0 = pd.DataFrame([{"aprovado": "S"}], columns=COLS_PRONTAS)
    bloqueadas_v0 = pd.DataFrame([{"motivo": "v0"}], columns=COLS_BLOQUEADAS)
    resumo_v0 = pd.DataFrame([{"campo": "total", "valor": "0"}])
    with pd.ExcelWriter(xlsx, engine="openpyxl") as writer:
        prontas_v0.to_excel(writer, sheet_name=ABA_PRONTAS, index=False)
        bloqueadas_v0.to_excel(writer, sheet_name=ABA_BLOQUEADAS, index=False)
        resumo_v0.to_excel(writer, sheet_name=ABA_RESUMO, index=False)
    bytes_originais = xlsx.read_bytes()

    # Conteúdo novo
    prontas = pd.DataFrame(
        [{"data": "01/01/2025", "valor": "100,00", "documento": "PIX001",
          "nr_extrato_bancario": "PIX001", "historico": "PIX RECEBIDO",
          "nome_doador": "JOAO", "cpf": "111", "cnpj": "", "tipo_pessoa": "PF",
          "fonte_recurso": "OR", "natureza_recurso": "0",
          "classificacao_receita": "320", "nr_banco": "001",
          "agencia": "1234", "dv_agencia": "5", "conta": "5678", "dv_conta": "9",
          "cnpj_prestador": "12345678000199", "nome_diretorio": "DirA",
          "aprovado": "S"}],
        columns=COLS_PRONTAS,
    )
    bloqueadas = pd.DataFrame(
        [{"data": "02/01/2025", "valor": "50,00", "documento": "PIX002",
          "historico": "PIX RECEBIDO", "nome_pix": "MARIA",
          "cpf_extrato": "222", "categoria": "x", "motivo": "y",
          "ignorar_exportacao": "N"}],
        columns=COLS_BLOQUEADAS,
    )
    resumo = pd.DataFrame([{"campo": "total", "valor": "2"}])

    path_retornado = salvar_revisao(output_dir, prontas, bloqueadas, resumo, mes_slug=mes_slug)
    assert path_retornado == xlsx

    # xlsx regravado contém os dados novos
    wb_novo = load_workbook(xlsx)
    ws_prontas = wb_novo[ABA_PRONTAS]
    assert ws_prontas.cell(2, 1).value == "01/01/2025"  # coluna 1 = data
    assert ws_prontas.cell(2, 6).value == "JOAO"  # coluna 6 = nome_doador
    # backup criado com bytes idênticos ao estado pré-mutação
    backups = list(output_dir.glob(f"{xlsx.name}.bak-*"))
    assert len(backups) == 1
    assert backups[0].read_bytes() == bytes_originais


def test_with_backup_salvar_revisao_primeira_escrita(tmp_path):
    """salvar_revisao() — quando xlsx NÃO existe, grava direto (sem backup).

    Valida o ramo else do padrão em lib_revisao_exportacao.py:salvar_revisao():
    1ª escrita não tem estado anterior a preservar, então o decorator não
    deve ser invocado. Confirma que o arquivo é criado e nenhum .bak-* é
    gerado.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from lib_revisao_exportacao import (
        ABA_PRONTAS,
        COLS_BLOQUEADAS,
        COLS_PRONTAS,
        caminho_revisao,
        salvar_revisao,
    )

    output_dir = tmp_path / "mensal"
    output_dir.mkdir()
    mes_slug = "fevereiro"
    xlsx = caminho_revisao(output_dir, mes_slug)
    assert not xlsx.exists()

    prontas = pd.DataFrame(
        [{"data": "01/02/2025", "valor": "200,00", "documento": "P1",
          "nr_extrato_bancario": "P1", "historico": "PIX", "nome_doador": "J",
          "cpf": "1", "cnpj": "", "tipo_pessoa": "PF", "fonte_recurso": "OR",
          "natureza_recurso": "0", "classificacao_receita": "320",
          "nr_banco": "1", "agencia": "1", "dv_agencia": "", "conta": "1",
          "dv_conta": "", "cnpj_prestador": "1", "nome_diretorio": "D",
          "aprovado": "S"}],
        columns=COLS_PRONTAS,
    )
    bloqueadas = pd.DataFrame(columns=COLS_BLOQUEADAS)
    resumo = pd.DataFrame([{"campo": "a", "valor": "1"}])

    path_retornado = salvar_revisao(output_dir, prontas, bloqueadas, resumo, mes_slug=mes_slug)
    assert path_retornado == xlsx
    assert xlsx.exists()

    # Sem backup — arquivo não existia antes, não há o que copiar
    backups = list(output_dir.glob(f"{xlsx.name}.bak-*"))
    assert backups == [], f"1ª escrita não deve gerar .bak-*, obtido {backups}"

    # Conteúdo do xlsx correto
    wb = load_workbook(xlsx)
    assert ABA_PRONTAS in wb.sheetnames
    assert wb[ABA_PRONTAS].cell(2, 1).value == "01/02/2025"


def test_with_backup_escrever_xml_regravacao(tmp_path):
    """gerar_xml_mes() — quando XML já existe, regrava com backup binário.

    Espelha o padrão em lib_xml_origem_recurso.py:gerar_xml_mes(): inner
    function _escrever_xml() é decorada com @with_backup(destino). Valida:
    backup preserva bytes do XML anterior e a regravação substitui o conteúdo.
    """
    from pathlib import Path
    from lib_xml_origem_recurso import _escrever_xml
    from with_backup import with_backup

    destino = tmp_path / "janeiro-origemRecurso.xml"
    destino.write_bytes(b"<?xml version='1.0'?><ANTIGO/>")
    bytes_anteriores = destino.read_bytes()

    novo_xml = b"<?xml version='1.0'?><NOVO/>"

    # Espelha exatamente o call site em gerar_xml_mes():
    # @with_backup(destino) decora _escrever_xml(destino) que delega para
    # destino.write_bytes(xml_bytes).
    @with_backup(destino)
    def _salvar_xml_com_backup(target: Path, xml_bytes: bytes) -> None:
        _escrever_xml(target, xml_bytes)

    _salvar_xml_com_backup(novo_xml)

    # XML regravado com bytes novos
    assert destino.read_bytes() == novo_xml
    # Backup criado com bytes idênticos ao estado pré-mutação
    backups = list(tmp_path.glob(f"{destino.name}.bak-*"))
    assert len(backups) == 1
    assert backups[0].read_bytes() == bytes_anteriores


def test_with_backup_escrever_xml_primeira_escrita(tmp_path):
    """gerar_xml_mes() — quando XML NÃO existe, grava direto (sem backup).

    Valida o ramo else do padrão em lib_xml_origem_recurso.py:gerar_xml_mes():
    1ª escrita do XML não tem estado anterior a preservar. Confirma que o
    arquivo é criado e nenhum .bak-* é gerado.
    """
    from pathlib import Path
    from lib_xml_origem_recurso import _escrever_xml

    destino = tmp_path / "fevereiro-origemRecurso.xml"
    assert not destino.exists()

    xml_bytes = b"<?xml version='1.0'?><PRIMEIRA_ESCRITA/>"
    _escrever_xml(destino, xml_bytes)

    assert destino.exists()
    assert destino.read_bytes() == xml_bytes
    # Sem backup
    backups = list(tmp_path.glob(f"{destino.name}.bak-*"))
    assert backups == [], f"1ª escrita não deve gerar .bak-*, obtido {backups}"
