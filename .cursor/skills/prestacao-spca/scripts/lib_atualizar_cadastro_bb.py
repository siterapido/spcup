#!/usr/bin/env python3
"""Inclui no cadastro TSE pessoas com CPF/CNPJ do extrato BB unificado (fora do cadastro)."""

from __future__ import annotations

import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from conciliar_doacoes import cpf_norm_compare, documento_extrato_bb_chaves, only_digits
from lib_export_mensal import _formatar_cpf_cnpj
from lib_paths import (
    MODELO_BB_UNIFICADO,
    carregar_meses,
    carregar_prestacao,
    nome_estado,
    normalizar_modelo_extrato,
    pasta_ano_prestacao,
    resolver_arquivo_mensal,
    resolver_cadastro,
)
from with_backup import with_backup


def _agora() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _ler_cadastro_sem_cabecalho(caminho: Path) -> list[tuple[str, str, str, str]]:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    linhas: list[tuple[str, str, str, str]] = []
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip()
        doc = str(row[1] or "").strip()
        tipo = str(row[2] or "Pessoa Física").strip() if len(row) > 2 else "Pessoa Física"
        status = str(row[3] or "Validado").strip() if len(row) > 3 else "Validado"
        linhas.append((nome, doc, tipo, status))
    wb.close()
    return linhas


def _chaves_cadastro(linhas: list[tuple[str, str, str, str]]) -> tuple[set[str], set[str]]:
    cpfs: set[str] = set()
    cnpjs: set[str] = set()
    for _, doc, _, _ in linhas:
        cpf_d, cnpj_d = documento_extrato_bb_chaves(doc)
        if cpf_d:
            cpfs.add(cpf_norm_compare(cpf_d))
        if cnpj_d:
            cnpjs.add(only_digits(cnpj_d))
    return cpfs, cnpjs


def _doc_ja_no_cadastro(raw: str, cpfs: set[str], cnpjs: set[str]) -> bool:
    cpf_d, cnpj_d = documento_extrato_bb_chaves(raw)
    if cpf_d and cpf_norm_compare(cpf_d) in cpfs:
        return True
    if cnpj_d and only_digits(cnpj_d) in cnpjs:
        return True
    return False


def _nome_pendencia(row: pd.Series) -> str:
    for col in ("Nome do Doador (PIX)", "Nome_PIX_original", "Nome do Doador"):
        val = str(row.get(col, "") or "").strip()
        if val:
            return val
    return ""


def _cpf_extrato_preenchido(row: pd.Series) -> bool:
    raw = str(row.get("CPF_extrato") or row.get("cpf_extrato") or "").strip()
    return bool(raw) and raw.lower() not in ("nan", "none")


def _pendencia_candidata_cadastro(row: pd.Series) -> bool:
    """Pendência com documento no extrato BB ainda ausente no cadastro TSE."""
    if not _cpf_extrato_preenchido(row):
        return False
    texto = f"{row.get('motivo', '')} {row.get('categoria', '')}".lower()
    if "fora do cadastro" in texto or "documento fora" in texto:
        return True
    if "cpf ausente" in texto:
        return True
    return False


def coletar_candidatos_extrato(
    raiz: Path,
    estado: str,
    ano: int,
    *,
    meses: list[str] | None = None,
    escopo: str | None = None,
) -> dict[str, dict[str, Any]]:
    """
    Únicos por chave de documento (CPF/CNPJ).
    Valor: nome (mais longo vence), doc formatado, tipo, meses vistos.
    """
    slugs = meses or list(carregar_meses().keys())
    candidatos: dict[str, dict[str, Any]] = {}

    prestacao = carregar_prestacao(raiz) or {
        "estado": estado,
        "estado_uf": "",
        "ano": ano,
        "escopo": escopo,
    }
    pasta_ano = pasta_ano_prestacao(raiz, prestacao)

    for slug in slugs:
        path = resolver_arquivo_mensal(pasta_ano, slug, "Pendencias_e_Inconsistencias.xlsx")
        if not path:
            continue
        df = pd.read_excel(path, sheet_name=0)
        if df.empty:
            continue
        for _, row in df.iterrows():
            if not _pendencia_candidata_cadastro(row):
                continue
            cpf_raw = str(row.get("CPF_extrato") or row.get("cpf_extrato") or "").strip()
            if not cpf_raw:
                continue
            cpf_d, cnpj_d = documento_extrato_bb_chaves(cpf_raw)
            if not cpf_d and not cnpj_d:
                continue
            chave = f"cpf:{cpf_norm_compare(cpf_d)}" if cpf_d else f"cnpj:{only_digits(cnpj_d)}"
            nome = _nome_pendencia(row)
            if not nome:
                continue
            doc_fmt = _formatar_cpf_cnpj(cpf_raw)
            tipo = "Pessoa Jurídica" if cnpj_d else "Pessoa Física"
            if chave not in candidatos or len(nome) > len(candidatos[chave]["nome"]):
                candidatos[chave] = {
                    "nome": nome,
                    "documento": doc_fmt,
                    "tipo": tipo,
                    "meses": set(),
                }
            candidatos[chave]["meses"].add(slug)

    return candidatos


def atualizar_cadastro_bb(
    raiz: Path,
    prestacao: dict[str, Any],
    *,
    meses: list[str] | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    modelo = normalizar_modelo_extrato(str(prestacao.get("modelo_extrato", "")))
    if modelo != MODELO_BB_UNIFICADO:
        raise ValueError(
            f"atualizar_cadastro_bb só para modelo bb_unificado (atual: {modelo!r})"
        )

    uf = prestacao["estado_uf"]
    estado = nome_estado(uf)
    ano = int(prestacao["ano"])
    base_prestacao = raiz / str(prestacao.get("base_prestacao", f"{estado}/Prestação de contas - {estado}"))
    cadastro_path = resolver_cadastro(raiz, uf, base_prestacao)

    linhas = _ler_cadastro_sem_cabecalho(cadastro_path)
    cpfs, cnpjs = _chaves_cadastro(linhas)
    escopo = str(prestacao.get("escopo") or "").strip() or None
    candidatos = coletar_candidatos_extrato(raiz, estado, ano, meses=meses, escopo=escopo)

    adicionar: list[tuple[str, str, str, str]] = []
    ignorados_cadastro: list[str] = []
    for chave, info in sorted(candidatos.items(), key=lambda x: x[1]["nome"]):
        if _doc_ja_no_cadastro(info["documento"], cpfs, cnpjs):
            ignorados_cadastro.append(chave)
            continue
        adicionar.append((info["nome"], info["documento"], info["tipo"], "Validado"))
        cpf_d, cnpj_d = documento_extrato_bb_chaves(info["documento"])
        if cpf_d:
            cpfs.add(cpf_norm_compare(cpf_d))
        if cnpj_d:
            cnpjs.add(only_digits(cnpj_d))

    backup_path: Path | None = None
    if adicionar and not dry_run:
        @with_backup(cadastro_path)
        def _gravar_adicoes(path, adicoes):
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            for nome, doc, tipo, status in adicoes:
                ws.append([nome, doc, tipo, status])
            wb.save(path)
            wb.close()

        _gravar_adicoes(adicionar)
        # o decorator deixou um backup .bak-<ts> ao lado do cadastro
        backups = sorted(cadastro_path.parent.glob(f"{cadastro_path.name}.bak-*"))
        if backups:
            backup_path = backups[-1]

    return {
        "estado": estado,
        "estado_uf": uf,
        "ano": ano,
        "modelo_extrato": modelo,
        "cadastro": str(cadastro_path),
        "backup": str(backup_path) if backup_path else None,
        "dry_run": dry_run,
        "candidatos_extrato": len(candidatos),
        "adicionados": len(adicionar),
        "ignorados_ja_no_cadastro": len(ignorados_cadastro),
        "linhas_cadastro_antes": len(linhas),
        "linhas_cadastro_depois": len(linhas) + len(adicionar),
        "pessoas_adicionadas": [
            {"nome": n, "documento": d, "tipo": t, "status": s}
            for n, d, t, s in adicionar[:20]
        ],
        "amostra_restante": max(0, len(adicionar) - 20),
        "atualizado_em": _agora(),
        "proximo_passo": (
            "Reconciliar meses sem novo NLM: processar_todos.py --estado ... --ano ... "
            "--forcar --pular-nlm (só se PDFs/NLM inalterados; recusa cadastro novo)."
        ),
    }
