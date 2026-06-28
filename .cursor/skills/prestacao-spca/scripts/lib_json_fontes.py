#!/usr/bin/env python3
"""Converte JSON (NLM / planilhado) em DataFrames para conciliar_doacoes."""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from lib_nlm import is_nlm_pix_recebido, origem_contem_pdf
from lib_paths import MODELO_BB_UNIFICADO, MODELO_CAIXA_1, normalizar_modelo_extrato


def _parse_hora_pix(hora: Any) -> tuple[int, int] | None:
    texto = str(hora or "").strip()
    if not texto:
        return None
    m = re.match(r"^(\d{1,2}):(\d{2})", texto)
    if m:
        return int(m.group(1)), int(m.group(2))
    if re.match(r"^\d{4}$", texto):
        return int(texto[:2]), int(texto[2:])
    if re.match(r"^\d{6}$", texto):
        return int(texto[:2]), int(texto[2:4])
    return None


def _ddhhmm_de_transacao(tx: dict[str, Any]) -> str:
    """DDHHMM a partir de data_pix+hora ou data+hora (extrato BB unificado)."""
    doc = re.sub(r"\D", "", str(tx.get("numero_documento") or ""))
    if len(doc) == 6:
        return doc
    data_raw = str(tx.get("data_pix") or tx.get("data") or "")[:10]
    hora = tx.get("hora")
    hm = _parse_hora_pix(hora)
    if not hm:
        return doc[-6:].zfill(6) if len(doc) >= 6 else ""
    dia = 0
    if data_raw:
        try:
            dia = datetime.strptime(data_raw, "%Y-%m-%d").day
        except ValueError:
            pass
    if dia:
        return f"{dia:02d}{hm[0]:02d}{hm[1]:02d}"
    return f"{hm[0]:02d}{hm[1]:02d}".zfill(6)


def _cpf_cnpj_de_transacao(tx: dict[str, Any]) -> str:
    for chave in (
        "cpf_cnpj_extrato",
        "cpf_cnpj",
        "documento_contraparte",
        "cpf",
        "cnpj",
    ):
        valor = str(tx.get(chave) or "").strip()
        if valor:
            return valor
    return ""


def carregar_cpfs_detalhe_pdf(pdf: Path) -> dict[str, str]:
    """Mapa DDHHMM → documento (CPF/CNPJ) a partir das linhas de detalhe do PDF BB."""
    import subprocess

    if not pdf.is_file():
        return {}
    try:
        text = subprocess.check_output(
            ["pdftotext", "-layout", str(pdf), "-"],
            text=True,
            errors="replace",
        )
    except (subprocess.CalledProcessError, FileNotFoundError):
        return {}

    out: dict[str, str] = {}
    pat = re.compile(
        r"^\s*(\d{2})/\d{2}\s+(\d{2}):(\d{2})\s+(\d{11,14})\s+",
        re.MULTILINE,
    )
    for m in pat.finditer(text):
        dd, hh, mi, doc = m.group(1), m.group(2), m.group(3), m.group(4)
        out[f"{int(dd):02d}{hh}{mi}"] = doc
    return out


def enriquecer_transacoes_bb_cpf(
    transacoes: list[dict[str, Any]],
    pdfs: list[str],
) -> int:
    """Preenche cpf_cnpj_extrato via linhas de detalhe do PDF quando NLM omitir."""
    cpf_map: dict[str, str] = {}
    for pdf_str in pdfs:
        cpf_map.update(carregar_cpfs_detalhe_pdf(Path(pdf_str)))

    if not cpf_map:
        return 0

    complementados = 0
    for tx in transacoes:
        if _cpf_cnpj_de_transacao(tx):
            continue
        ddhhmm = _ddhhmm_de_transacao(tx)
        doc = cpf_map.get(ddhhmm)
        if doc:
            tx["cpf_cnpj_extrato"] = doc
            complementados += 1
    return complementados


def split_bb_unificado(transacoes: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """
    Extrato BB unificado (Santa Catarina): uma linha com histórico Pix - Recebido,
    valor, data balancete e detalhe (data_pix, hora, CPF, nome).
  Gera extrato_total + extrato_pix para o pipeline existente.
    """
    from conciliar_doacoes import is_cred_pix_recebido, normalize_historico

    total: list[dict[str, Any]] = []
    pix: list[dict[str, Any]] = []
    for tx in transacoes:
        if not is_nlm_pix_recebido(tx):
            continue
        hist = str(tx.get("tipo") or "Pix - Recebido").strip()
        hist_norm = normalize_historico(hist)
        if not is_cred_pix_recebido(hist_norm):
            continue
        valor = tx.get("valor")
        data_mov = str(tx.get("data") or "")[:10]
        data_pix = str(tx.get("data_pix") or data_mov)[:10]
        hora = str(tx.get("hora") or "").strip()
        nome = str(tx.get("remetente_destinatario") or "").strip()
        doc_cpf = _cpf_cnpj_de_transacao(tx)
        doc = _ddhhmm_de_transacao(tx)
        total.append(
            {
                "Data": _data_br(data_mov),
                "Valor": _valor_br(valor),
                "Documento": doc,
                "Histórico": hist,
            }
        )
        if nome and hora:
            linha_pix: dict[str, Any] = {
                "Data": _data_br(data_pix),
                "Hora": hora if ":" in hora else f"{hora[:2]}:{hora[2:4]}:00",
                "Valor": _valor_br(valor),
                "Remetente/Destinatario": nome,
            }
            if doc_cpf:
                linha_pix["CPF/CNPJ"] = doc_cpf
            pix.append(linha_pix)
    return total, pix


def _data_br(iso: str | None) -> str:
    if not iso:
        return ""
    texto = str(iso).strip()[:10]
    try:
        dt = datetime.strptime(texto, "%Y-%m-%d")
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        return texto


def _valor_br(valor: Any) -> str:
    try:
        v = round(float(valor), 2)
    except (TypeError, ValueError):
        return str(valor or "")
    inteiro, frac = f"{v:.2f}".split(".")
    return f"{inteiro},{frac}"


def _parse_data_planilha(valor: Any) -> str:
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor or "").strip()
    if "/" in texto:
        partes = texto.split("/")
        if len(partes) >= 3:
            dia, mes, ano = partes[0], partes[1], partes[2][:4]
            return f"{ano}-{mes.zfill(2)}-{dia.zfill(2)}"
    return texto[:10]


def _parse_valor_planilha(valor: Any) -> float:
    if isinstance(valor, (int, float)):
        return round(float(valor), 2)
    texto = str(valor or "").replace("R$", "").replace("\xa0", "").replace(",", ".").strip()
    try:
        return round(float(re.sub(r"[^0-9.]", "", texto) or 0), 2)
    except ValueError:
        return 0.0


def _primeiro_xlsx(pasta: Path) -> Path | None:
    arquivos = sorted(pasta.glob("*.xlsx"))
    return arquivos[0] if arquivos else None


def ler_planilhado_pix(pasta: Path) -> list[dict[str, Any]]:
    if pasta.is_file() and pasta.suffix.lower() in {".xlsx", ".xls"}:
        xlsx = pasta
    else:
        xlsx = _primeiro_xlsx(pasta)
    if not xlsx:
        return []

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    header = [str(c or "").strip() for c in next(ws.iter_rows(max_row=1, values_only=True))]
    col_tipo = next((i for i, h in enumerate(header) if "pix" in h.lower()), 2)
    col_nome = next(
        (i for i, h in enumerate(header) if "remetente" in h.lower() or "destinat" in h.lower()),
        4,
    )
    col_valor = next((i for i, h in enumerate(header) if "valor" in h.lower()), 5)
    col_data = next((i for i, h in enumerate(header) if h.lower() == "data"), 0)
    col_hora = next((i for i, h in enumerate(header) if "hora" in h.lower()), None)

    registros: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col_nome:
            continue
        tipo_pix = str(row[col_tipo] or "").strip()
        if tipo_pix.lower() != "recebido":
            continue
        nome = str(row[col_nome] or "").strip()
        if not nome:
            continue
        data_iso = _parse_data_planilha(row[col_data]) if col_data is not None else ""
        hora = (
            str(row[col_hora] or "").strip()[:8]
            if col_hora is not None and len(row) > col_hora
            else ""
        )
        valor_num = _parse_valor_planilha(row[col_valor] if len(row) > col_valor else None)
        registros.append(
            {
                "data": data_iso,
                "hora": hora,
                "valor": valor_num,
                "nome": nome,
                "arquivo": xlsx.name,
            }
        )
    wb.close()
    return registros


def nlm_para_extrato_total(
    transacoes: list[dict[str, Any]],
    pdfs_total: list[str],
) -> list[dict[str, Any]]:
    linhas: list[dict[str, Any]] = []
    for tx in transacoes:
        origem = str(tx.get("origem_arquivo") or "")
        if not origem_contem_pdf(origem, pdfs_total):
            continue
        linhas.append(
            {
                "Data": _data_br(tx.get("data")),
                "Valor": _valor_br(tx.get("valor")),
                "Documento": str(tx.get("numero_documento") or "").strip(),
                "Histórico": str(tx.get("tipo") or "").strip(),
            }
        )
    return linhas


def nlm_para_extrato_pix(
    transacoes: list[dict[str, Any]],
    pdfs_pix: list[str],
) -> list[dict[str, Any]]:
    linhas: list[dict[str, Any]] = []
    vistos: set[tuple[str, str, str, str]] = set()
    for tx in transacoes:
        origem = str(tx.get("origem_arquivo") or "")
        if not origem_contem_pdf(origem, pdfs_pix):
            continue
        if not is_nlm_pix_recebido(tx):
            continue
        nome = str(tx.get("remetente_destinatario") or "").strip()
        if not nome:
            continue
        data_iso = str(tx.get("data") or "")[:10]
        hora = str(tx.get("hora") or "").strip()
        if not hora:
            doc = re.sub(r"\D", "", str(tx.get("numero_documento") or ""))
            if len(doc) >= 4:
                hhmm = doc[-4:]
                hora = f"{hhmm[:2]}:{hhmm[2:4]}:00"
        # Deduplicar por (data, hora em minuto, valor, nome)
        hora_min = hora[:5] if ":" in hora else hora[:4]
        data_br = _data_br(data_iso)
        valor_br = _valor_br(tx.get("valor"))
        chave = (data_br, hora_min, valor_br, nome.upper())
        if chave in vistos:
            continue
        vistos.add(chave)
        linha: dict[str, Any] = {
            "Data": data_br,
            "Hora": hora,
            "Valor": valor_br,
            "Remetente/Destinatario": nome,
        }
        cpf_cnpj = _cpf_cnpj_de_transacao(tx)
        if cpf_cnpj:
            linha["CPF/CNPJ"] = cpf_cnpj
        linhas.append(linha)
    return linhas


def planilhado_para_extrato_pix(registros: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "Data": _data_br(reg.get("data")),
            "Hora": reg.get("hora") or "",
            "Valor": _valor_br(reg.get("valor")),
            "Remetente/Destinatario": reg.get("nome") or "",
        }
        for reg in registros
    ]


def _ddhhmm_de_hora(data_br: str, hora: str) -> str:
    """Gera documento DDHHMM a partir de data DD/MM/YYYY e hora HH:MM:SS."""
    hm = _parse_hora_pix(hora)
    if not hm:
        return ""
    try:
        dia = datetime.strptime(data_br, "%d/%m/%Y").day
        return f"{dia:02d}{hm[0]:02d}{hm[1]:02d}"
    except ValueError:
        return f"{hm[0]:02d}{hm[1]:02d}".zfill(6)


def extrato_total_a_partir_do_pix(extrato_pix: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Fallback Caixa: quando NLM não identifica CRED PIX no total, deriva do PIX."""
    total: list[dict[str, Any]] = []
    for pix in extrato_pix:
        data_br = pix.get("Data") or ""
        hora = pix.get("Hora") or ""
        valor = pix.get("Valor") or ""
        doc = _ddhhmm_de_hora(data_br, hora)
        total.append(
            {
                "Data": data_br,
                "Valor": valor,
                "Documento": doc,
                "Histórico": "CRED PIX",
            }
        )
    return total


def _tem_cred_pix(extrato_total: list[dict[str, Any]]) -> bool:
    for linha in extrato_total:
        hist = str(linha.get("Histórico") or "").upper()
        if "CRED PIX" in hist:
            return True
    return False


def xlsx_pessoas_para_json(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    header = [str(c or "").strip().lower() for c in rows[0]]
    idx_nome = next((i for i, h in enumerate(header) if "nome" in h or "raz" in h), 0)
    idx_doc = next((i for i, h in enumerate(header) if "cpf" in h or "cnpj" in h or h == "documento"), 1)
    idx_tipo = next((i for i, h in enumerate(header) if "tipo" in h), 2)
    idx_status = next((i for i, h in enumerate(header) if "status" in h or "valid" in h), 3)

    pessoas: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or not row[idx_nome]:
            continue
        pessoas.append(
            {
                "nome": str(row[idx_nome]).strip(),
                "documento": str(row[idx_doc] or "").strip() if idx_doc < len(row) else "",
                "tipo": str(row[idx_tipo] or "").strip() if idx_tipo < len(row) else "",
                "status": str(row[idx_status] or "Validado").strip()
                if idx_status < len(row)
                else "Validado",
            }
        )
    return pessoas


def pessoas_json_para_df(pessoas: list[dict[str, Any]]) -> pd.DataFrame:
    colunas = ["Nome", "CPF/CNPJ", "Tipo de Pessoa", "Status"]
    linhas = [
        {
            "Nome": p.get("nome") or "",
            "CPF/CNPJ": p.get("documento") or "",
            "Tipo de Pessoa": p.get("tipo") or "",
            "Status": p.get("status") or "Validado",
        }
        for p in pessoas
        if p.get("nome")
    ]
    if not linhas:
        return pd.DataFrame(columns=colunas)
    return pd.DataFrame(linhas)


def salvar_fontes_json(
    cache_dir: Path,
    *,
    extrato_total: list[dict[str, Any]],
    extrato_pix: list[dict[str, Any]],
    pessoas: list[dict[str, Any]],
    meta: dict[str, Any] | None = None,
) -> Path:
    cache_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "versao": 1,
        "extrato_total": extrato_total,
        "extrato_pix": extrato_pix,
        "pessoas": pessoas,
        "meta": meta or {},
    }
    destino = cache_dir / "fontes.json"
    destino.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return destino


def carregar_fontes_json(cache_dir: Path) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    path = cache_dir / "fontes.json"
    if not path.is_file():
        raise FileNotFoundError(f"fontes.json ausente em {cache_dir}. Execute extração NLM primeiro.")
    dados = json.loads(path.read_text(encoding="utf-8"))
    df_total = pd.DataFrame(dados.get("extrato_total", []), dtype=str)
    df_pix = pd.DataFrame(dados.get("extrato_pix", []), dtype=str)
    df_pessoas = pessoas_json_para_df(dados.get("pessoas", []))
    return df_total, df_pix, df_pessoas


def montar_fontes_json(
    paths: dict[str, Any],
    *,
    nlm_meta: dict[str, Any] | None = None,
) -> Path:
    """Monta fontes.json a partir de NLM, planilhado e/ou cadastro XLSX."""
    cache_dir = Path(paths["cache_dir"])
    nlm_path = cache_dir / "nlm_transacoes.json"
    transacoes: list[dict[str, Any]] = []
    if nlm_path.is_file():
        transacoes = json.loads(nlm_path.read_text(encoding="utf-8")).get("transacoes", [])

    pdfs_total = [Path(p).name for p in paths["pdfs_total"]]
    pdfs_pix = [Path(p).name for p in paths["pdfs_pix"]]
    modelo = normalizar_modelo_extrato(
        paths.get("modelo_extrato") or (nlm_meta or {}).get("modelo_extrato")
    )

    if modelo == MODELO_BB_UNIFICADO and transacoes:
        pdfs_paths = [str(p) for p in paths.get("pdfs_total") or []]
        enriquecer_transacoes_bb_cpf(transacoes, pdfs_paths)
        extrato_total, extrato_pix = split_bb_unificado(transacoes)
        fonte_pix = "nlm_bb_unificado"
    else:
        extrato_total = nlm_para_extrato_total(transacoes, pdfs_total)
        planilhado_regs: list[dict[str, Any]] = []
        pasta_plan = paths.get("pasta_planilhado")
        if pasta_plan:
            plan_path = Path(pasta_plan)
            if plan_path.is_dir() or (plan_path.is_file() and plan_path.suffix == ".xlsx"):
                planilhado_regs = ler_planilhado_pix(plan_path)

        # PIX: NLM (PDF) tem prioridade — planilhado só se NLM não retornar linhas
        extrato_pix = nlm_para_extrato_pix(transacoes, pdfs_pix)
        if extrato_pix:
            fonte_pix = "nlm"
        elif planilhado_regs:
            extrato_pix = planilhado_para_extrato_pix(planilhado_regs)
            fonte_pix = "planilhado"
        else:
            extrato_pix = []
            fonte_pix = "ausente"

        # Fallback Caixa: se o total não tem CRED PIX mas o PIX tem entradas,
        # derivar o total a partir do PIX (NLM confundiu histórico do extrato total).
        if modelo == MODELO_CAIXA_1 and extrato_pix and not _tem_cred_pix(extrato_total):
            extrato_total = extrato_total_a_partir_do_pix(extrato_pix)

    if not extrato_total:
        raise ValueError("extrato_total vazio após NLM — verifique PDFs do extrato total.")

    if not extrato_pix:
        raise ValueError("extrato_pix vazio — forneça Planilhado/*.xlsx ou PDF PIX para NLM.")

    pessoas: list[dict[str, Any]] = []
    if paths.get("path_pessoas"):
        pessoas = xlsx_pessoas_para_json(Path(paths["path_pessoas"]))
        fonte_pessoas = "cadastro_xlsx"
    else:
        nlm_pessoas = cache_dir / "nlm_pessoas.json"
        if nlm_pessoas.is_file():
            pessoas = json.loads(nlm_pessoas.read_text(encoding="utf-8")).get("pessoas", [])
            fonte_pessoas = "nlm_pessoas"
        else:
            raise FileNotFoundError(
                "Cadastro ausente. Coloque pessoas.xlsx no cadastro ou pessoas.pdf para NLM."
            )

    meta = {
        **(nlm_meta or {}),
        "modelo_extrato": normalizar_modelo_extrato(paths.get("modelo_extrato")) or MODELO_CAIXA_1,
        "fonte_pix": fonte_pix,
        "fonte_pessoas": fonte_pessoas,
        "linhas_total": len(extrato_total),
        "linhas_pix": len(extrato_pix),
        "linhas_pessoas": len(pessoas),
    }
    return salvar_fontes_json(
        cache_dir,
        extrato_total=extrato_total,
        extrato_pix=extrato_pix,
        pessoas=pessoas,
        meta=meta,
    )
