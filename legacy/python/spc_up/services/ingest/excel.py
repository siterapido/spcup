"""Excel spreadsheet parsing for bank transactions."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from spc_up.models.entities import MovimentacaoDirecao

_REQUIRED_COLUMNS = frozenset({"data", "valor", "descricao"})
_OPTIONAL_COLUMNS = frozenset({"tipo"})
_VALID_HEADERS = _REQUIRED_COLUMNS | _OPTIONAL_COLUMNS

_TIPO_ENTRADA = frozenset({"C", "CREDITO", "CRÉDITO", "ENTRADA"})
_TIPO_SAIDA = frozenset({"D", "DEBITO", "DÉBITO", "SAIDA", "SAÍDA"})


def _normalize_header(value: Any) -> str | None:
    if value is None:
        return None
    key = str(value).strip().lower()
    if key in _VALID_HEADERS:
        return key
    return None


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Data inválida: {value!r}")


def _parse_decimal(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        raise ValueError(f"Valor inválido: {value!r}")
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = str(value).strip()
    if not text:
        raise ValueError("Valor inválido: vazio")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Valor inválido: {value!r}") from exc


def _direction_from_tipo(tipo: Any, amount: Decimal) -> tuple[str, Decimal]:
    if tipo is not None and str(tipo).strip():
        token = str(tipo).strip().upper()
        if token in _TIPO_ENTRADA:
            return MovimentacaoDirecao.ENTRADA.value, abs(amount)
        if token in _TIPO_SAIDA:
            return MovimentacaoDirecao.SAIDA.value, abs(amount)
        raise ValueError(f"Tipo inválido: {tipo!r}")
    if amount >= 0:
        return MovimentacaoDirecao.ENTRADA.value, abs(amount)
    return MovimentacaoDirecao.SAIDA.value, abs(amount)


def _row_is_empty(values: dict[str, Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values.values())


def parse_excel(path: str | Path) -> list[dict[str, Any]]:
    """Parse an Excel file into normalized transaction rows (same shape as OFX ingest)."""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            return []

        header_row: list[Any] | None = None
        column_index: dict[str, int] = {}
        rows: list[dict[str, Any]] = []

        for row in sheet.iter_rows(values_only=True):
            if header_row is None:
                header_row = list(row)
                for index, cell in enumerate(header_row):
                    name = _normalize_header(cell)
                    if name is not None:
                        column_index[name] = index
                missing = _REQUIRED_COLUMNS - column_index.keys()
                if missing:
                    labels = ", ".join(sorted(missing))
                    raise ValueError(f"Colunas obrigatórias ausentes: {labels}")
                continue

            values = {
                name: row[column_index[name]] if column_index[name] < len(row) else None
                for name in column_index
            }
            if _row_is_empty(values):
                continue

            amount = _parse_decimal(values["valor"])
            direcao, valor = _direction_from_tipo(values.get("tipo"), amount)
            descricao = str(values["descricao"] or "").strip()

            rows.append(
                {
                    "data_movimento": _parse_date(values["data"]),
                    "valor": valor,
                    "descricao_raw": descricao,
                    "direcao": direcao,
                    "nr_extrato_bancario": None,
                }
            )
        return rows
    finally:
        workbook.close()
